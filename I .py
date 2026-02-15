from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import LineChart, PieChart, Reference

# -----------------------------
# CREATE WORKBOOK
# -----------------------------
wb = Workbook()

# ============================================================
# SHEET 1  DAILY TRACKER
# ============================================================

ws = wb.active
ws.title = "DAILY TRACKER"

# Header row
ws["A1"] = "Habit"
for day in range(1, 32):
    ws.cell(row=1, column=day+1).value = day

ws["AG1"] = "Total"
ws["AH1"] = "Target"
ws["AI1"] = "%"

ws["A1"].font = Font(bold=True)

# Habit list
habits = [
    "Wake up at 5AM ",
    "No Snoozing ",
    "Drink 3L Water ",
    "Strength Training ",
    "Stretching ",
    "3D Practice ",
    "Coding Practice ",
    "Meditation 10m ",
    "Limit Social Media ",
    "No Alcohol "
]

for i, habit in enumerate(habits, start=2):
    ws[f"A{i}"] = habit
    ws[f"AG{i}"] = f"=COUNTIF(B{i}:AF{i},TRUE)"
    ws[f"AH{i}"] = 30
    ws[f"AI{i}"] = f"=AG{i}/AH{i}*100"

# -----------------------------
# DAILY % ROW
# -----------------------------
daily_percent_row = 13
ws[f"A{daily_percent_row}"] = "Daily %"

for col in range(2, 33):
    ws.cell(row=daily_percent_row, column=col).value = \
        f"=COUNTIF({chr(64+col)}2:{chr(64+col)}11,TRUE)/COUNTA($A$2:$A$11)*100"

# ============================================================
# SHEET 2  DASHBOARD
# ============================================================

dash = wb.create_sheet("DASHBOARD")

dash["A1"] = "GLOBAL PROGRESS"
dash["A1"].font = Font(size=14, bold=True)

dash["A3"] = "Completed"
dash["B3"] = "=SUM('DAILY TRACKER'!AG2:AG11)"

dash["A4"] = "Goal"
dash["B4"] = "=SUM('DAILY TRACKER'!AH2:AH11)"

dash["A5"] = "Left"
dash["B5"] = "=B4-B3"

dash["A6"] = "Completion %"
dash["B6"] = "=B3/B4*100"

# -----------------------------
# WEEKLY CALCULATIONS
# -----------------------------
dash["A8"] = "WEEKLY PROGRESS"

for i in range(4):
    row = 9 + i
    start_col = 2 + (i * 7)
    end_col = start_col + 6
    start_letter = chr(64 + start_col)
    end_letter = chr(64 + end_col)

    dash[f"A{row}"] = f"Week {i+1}"
    dash[f"B{row}"] = \
        f"=SUM('DAILY TRACKER'!{start_letter}2:{end_letter}11)"
    dash[f"C{row}"] = \
        f"=7*COUNTA('DAILY TRACKER'!A2:A11)"
    dash[f"D{row}"] = f"=B{row}/C{row}*100"

# -----------------------------
# PERFORMANCE RATING
# -----------------------------
dash["A14"] = "Performance Rating"
dash["B14"] = '=IF(B6>=90,"Elite ",IF(B6>=75,"Strong ",IF(B6>=60,"Moderate ","Needs Reset ")))'

# ============================================================
# CHARTS
# ============================================================

# Line Chart  Daily %
line = LineChart()
line.title = "Daily Progress %"
line.y_axis.title = "Percentage"
line.x_axis.title = "Day"

data = Reference(ws, min_col=2, max_col=32,
                 min_row=daily_percent_row,
                 max_row=daily_percent_row)

line.add_data(data, titles_from_data=False)
dash.add_chart(line, "F2")

# Pie Chart  Overall
pie = PieChart()
pie.title = "Overall Completion"

pie_data = Reference(dash, min_col=2, min_row=3, max_row=5)
pie_labels = Reference(dash, min_col=1, min_row=3, max_row=5)

pie.add_data(pie_data, titles_from_data=False)
pie.set_categories(pie_labels)

dash.add_chart(pie, "F18")

# ============================================================
# SAVE FILE
# ============================================================

wb.save("Ultimate_Habit_Dashboard.xlsx")

print("Ultimate Habit Dashboard Created Successfully!")